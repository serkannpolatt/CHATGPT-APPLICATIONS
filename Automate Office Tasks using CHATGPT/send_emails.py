import os
import openpyxl
import win32com.client as win32


# Get the current working directory
# Geçerli çalışma dizinini al
cwd = os.getcwd()

# Load the Excel workbook
# Excel çalışma kitabını yükleyin
workbook = openpyxl.load_workbook(os.path.join(cwd, "01-Distribute Excel Files with Outlook\Financial_Data.xlsx"))

# Select the sheet
# Sayfayı seçin
sheet = workbook["Email_List"]

# Get the Outlook application object
# Outlook uygulama nesnesini alın
outlook = win32.Dispatch('outlook.application')

# Iterate through the rows in the sheet
# Sayfadaki satırları yineleyin
for i in range(2, sheet.max_row + 1):

    # Get the attachment file name
    # Ek dosya adını al
    attachment = sheet.cell(row=i, column=1).value
    attachment_path = os.path.join(cwd, "Attachments", attachment)
    if not os.path.exists(attachment_path):
        print(f"Attachment {attachment} does not exist")
        continue

    # Get the recipient name
    # Alıcı adını al
    recipient_name = sheet.cell(row=i, column=2).value

    # Get the recipient email address
    # Alıcının e-posta adresini alın
    recipient_email = sheet.cell(row=i, column=3).value

    # Get the CC email address
    # CC e-posta adresini alın
    cc_email = sheet.cell(row=i, column=4).value

    # Create a new email
    # Yeni bir e-posta oluştur
    mail = outlook.CreateItem(0)

    # Set the recipient and CC email addresses
    # Alıcıyı ve CC e-posta adreslerini ayarlayın
    mail.To = recipient_email
    mail.CC = cc_email

    # Set the email subject
    # E-posta konusunu ayarlayın
    mail.Subject = f"Financial Data: {attachment}"

    # Set the email text
    # E-posta metnini ayarlayın
    mail.Body = f"Dear {recipient_name},\n\nPlease find the attached financial data for {attachment}.\n\nBest regards,\nYour Name"

    # Add the attachment
    # Eki ekle
    mail.Attachments.Add(attachment_path)

    # Open the email in Outlook
    # E-postayı Outlook'ta açın
    mail.Display()
    
# close all opened objects
# açılan tüm nesneleri kapat
workbook.close()
